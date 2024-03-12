# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import openpyxl
import os
import yaml

# 定义一个函数，将一行数据转换为字典
def row_to_dict(row):
    # 原胚类型的值定义材料和描述的映射
    Quality_mapping = {
        0: "&7劣质",
        1: "&f普通",
        2: "&a精良",
        3: "&3稀有",
        4: "&5史诗",
        5: "&6传说"
    }
    Tpye_mapping = {
        0: "&2&l匕首 &7&l原胚'",
        1: "&2&l轻剑 &7&l原胚'",
        2: "&2&l短锤 &7&l原胚'",
        3: "&2&l刀 &7&l原胚'",
        4: "&2&l战斧 &7&l原胚'",
        5: "&2&l月刃 &7&l原胚'",
        6: "&2&l长枪 &7&l原胚'",
        7: "&2&l重剑 &7&l原胚'",
        8: "&2&l长弓 &7&l原胚'",
        9: "&2&l短弓 &7&l原胚'",
        10: "&2&l枪械 &7&l原胚'",
        11: "&2&l长杖 &7&l原胚'",
        12: "&2&l短杖 &7&l原胚'",
        13: "&2&l头盔 &7&l原胚'",
        14: "&2&l胸甲 &7&l原胚'",
        15: "&2&l护腿 &7&l原胚'",
        16: "&2&l靴子 &7&l原胚'",

    }
    DamageTpye_mapping = {
        0: "!!物理倍率:1",
        1: "!!能量倍率:1",
    }


    # 创建字典
    data_dict = {
        row[1]: {
           "material": "GHAST_TEAR",
            "title": f"&5&l{row[1]}(原胚)",
            "lore": [
                f"-public/test/{row[1]}原胚",
                Quality_mapping[row[2]],
                Tpye_mapping[row[3]],
                "~preset_jichu",
                "&7&l用于锻造时:",

            ],
            "meta": {
                "unbreakable": False,
                "looksEnchanted": False
            }
        }
    }
    # 遍历G列到AI列，记录下所有不为空的数据
    if row[4] is not None:
        data_dict[row[1]]["lore"].append(f"&b{DamageTpye_mapping[row[4]]}")
    for i in range(5, 15):
        if row[i] is not None:
            data_dict[row[1]]["lore"].append(f"&b{sheet.cell(row=1, column=i + 1).value}转化 &7{row[i]:.0%}")

    data_dict[row[1]]["lore"].append("'")
    data_dict[row[1]]["lore"].append("&7&l装备被动:")
    for i in range(15, 24):
        if row[i] is not None:
            if (i % 3 == 0):
                data_dict[row[1]]["lore"].append(f"!!&b{row[i]}")
            if (i % 3 == 1):
                data_dict[row[1]]["lore"].append(f"&b{row[i]}")
            if (i % 3 == 2):
                chunks = [row[i][j:j + 13] for j in range(0, len(row[i]), 13)]
                # 在 lore 中输出每一组
                for chunk in chunks:
                    data_dict[row[1]]["lore"].append(f"&7{chunk}")

    # 简介的制作
    data_dict[row[1]]["lore"].append("\",")
    data_dict[row[1]]["lore"].append("~!preset_jianjie,")
    if row[-1] is not None:
        chunks = [row[-1][i:i + 13] for i in range(0, len(row[-1]), 13)]
        # 在 lore 中输出每一组
        for chunk in chunks:
            data_dict[row[1]]["lore"].append(f"&7{chunk}")
        data_dict[row[1]]["lore"].append("~end")
    return data_dict


# 加载工作簿
wb = openpyxl.load_workbook("H:\\fantasyrealm\\文案\\绘境\\总文档\\1040000.xlsx")

# 获取第一个工作表
sheet = wb.active

# 遍历工作表中的每一行
# 遍历工作表中的每一行
for row in sheet.iter_rows(min_row=3,min_col=1, values_only=True):
    # 将行数据转换为字典
    data_dict = row_to_dict(row)

    # 将字典转换为YAML字符串
    yaml_str = yaml.dump(data_dict, allow_unicode=True)

    # 获取main.py所在的目录
    base_dir = os.path.dirname(os.path.abspath(__file__))

    # 创建锻材文件夹的路径
    folder_path = os.path.join(base_dir, "原胚")

    # 如果锻材文件夹不存在，则创建
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # 将YAML字符串写入锻材文件夹中的文件
    with open(os.path.join(folder_path, f"{row[1]}.yml"), "w", encoding='utf-8') as f:
        f.write(yaml_str)
print("YAML文件已成功创建。")