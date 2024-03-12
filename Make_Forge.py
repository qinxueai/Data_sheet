# This is a sample Python script.
# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import openpyxl
import os
import yaml
# 定义一个函数，将一行数据转换为字典
def row_to_dict(row):
    # 锻材类型的值定义材料和描述的映射
    material1_mapping = {
        0: "&c&l矿物质 &7&l材料",
        1: "&5&l灵质 &7&l材料",
        2: "&2&l木质 &7&l材料",
        3: "&4&l生物质 &7&l材料"
    }
    material2_mapping = {
        0: "-矿物质 材料",
        1: "-灵质 材料",
        2: "-木质 材料",
        3: "-生物质 材料"
    }
    star_mapping = {
        1: "⭐",
        2: "⭐⭐",
        3: "⭐⭐⭐",
        4: "⭐⭐⭐⭐",
        5: "⭐⭐⭐⭐⭐",
        6: "⭐⭐⭐⭐⭐⭐",
        7: "⭐⭐⭐⭐⭐⭐⭐",
    }
    # 创建字典
    data_dict = {
        row[1]: {
           "material": "GHAST_TEAR",
            "title": f"&5&l{row[1]}",
            "lore": [
                f"-",
                material1_mapping[row[2]],
                "&7&l" + star_mapping[row[3]],
                material2_mapping[row[2]],
                "~preset_jichu",
                "&7&l用于锻造时:",

            ],
            "meta": {
                "unbreakable": False,
                "looksEnchanted": False
            }
        }
    }
    # 主属性的制作
    if int(row[2]) == 0:
        for i in range(36, 46):
            if row[i] is not None:
                data_dict[row[1]]["lore"].append(f"&b{sheet.cell(row=1, column=i + 1).value} &7&l+{row[i]}")
            else:
                print(f"{row[1]}矿物质属性不全诶？")
    else:
        # 初始化一个空字符串
        attributes_str = "&b{select<->"

        # 遍历G列到AI列，记录下所有不为空的数据
        for i in range(7, 36):
            if row[i] is not None:
                # 检查单元格数据是否以百分号结尾
                ends_with_percent = str(row[i]).endswith('%')
                column_name = sheet.cell(row=1, column=i + 1).value.rstrip('%') if ends_with_percent else sheet.cell(
                    row=1, column=i + 1).value
                value = row[i].rstrip('%') if ends_with_percent else row[i]
                attributes_str += f"{column_name} &7+[rand<=>{value}]" + ("(%)," if ends_with_percent else ",")
        # 移除最后一个逗号并添加"}"
        attributes_str = attributes_str.rstrip(',') + "}"
        # 将属性字符串添加到列表中
        data_dict[row[1]]["lore"].append(attributes_str)

        # 副属性的制作
    def write_attribute(list):
        attributes_str = "&b{select<->"
        for i in range(0, len(list)):
            if str(attribute_dict[(list[i], row[3])]).endswith('%'):
                if list[i].endswith('%'):
                    attributes_str += f"{list[i].rstrip('%')} &7+[rand<=>0-{attribute_dict[(list[i], row[3])].rstrip('%')}](%),"
                else:
                    attributes_str += f"{list[i]} &7+[rand<=>0-{attribute_dict[(list[i], row[3])].rstrip('%')}]%,"
            else:
                attributes_str += f"{list[i]} &7+[rand<=>0-{attribute_dict[(list[i], row[3])]}],"
        # 移除最后一个逗号并添加"}"
        attributes_str = attributes_str.rstrip(',') + "}"
        data_dict[row[1]]["lore"].append(attributes_str)

    # 第一副属性制作
    if row[6] is None:
        write_attribute(attribute_list)
    if row[6] is not None:
        Secondary_attribute_list = row[6].split(' ')
        write_attribute(Secondary_attribute_list)
            # 副属性的个数
        # Add missing import statement for openpyxl
    if (row[3] == 1 or row[3] == 2):
        n = 0
    elif (row[3] == 3 or row[3] == 4 or row[3] == 5):
        n = 1
    elif (row[3] == 6 or row[3] == 7):
        n = 2
    for j in range(n):
        write_attribute(attribute_list)


    # 简介的制作
    data_dict[row[1]]["lore"].append("")
    data_dict[row[1]]["lore"].append("~!preset_jianjie")
    chunks = [row[-1][i:i + 13] for i in range(0, len(row[-1]), 13)]
    # 在 lore 中输出每一组
    for chunk in chunks:
        data_dict[row[1]]["lore"].append(f"&7{chunk}")
    data_dict[row[1]]["lore"].append("~end")


    return data_dict
# 加载副属性工作簿
wb = openpyxl.load_workbook("H:\\fantasyrealm\\文案\\绘境\\总文档\\副属性表.xlsx", read_only=True, data_only=True)

# 选择第一个工作表
sheet = wb.active

# 初始化一个空字典和一个空列表
attribute_dict = {}
attribute_list = []

# 遍历工作表中的每一行
for row in range(2, sheet.max_row + 1):
    # 第一列是属性名
    attribute = sheet.cell(row=row, column=1).value

    attribute_list.append(attribute)
    # 遍历行中的剩余列
    for col in range(2, sheet.max_column + 1):
        # 键是一个元组，由属性和列号组成（列号对应星级）
        key = (attribute, col - 1)

        # 将键值对添加到字典中
        attribute_dict[key] = sheet.cell(row=row, column=col).value


# 定义副属性的属性
# 加载工作簿
wb = openpyxl.load_workbook("H:\\fantasyrealm\\文案\\绘境\\总文档\\1080000.xlsx")

# 获取第一个工作表
sheet = wb.active

# 遍历工作表中的每一行
for row in sheet.iter_rows(min_row=3,min_col=1, values_only=True):
    # 将行数据转换为字典
    data_dict = row_to_dict(row)
    # 将字典转换为YAML字符串
    yaml_str = yaml.dump(data_dict, allow_unicode=True)

    # 获取main.py所在的目录
    base_dir = os.path.dirname(os.path.abspath(__file__))

    # 创建锻材文件夹的路径
    folder_path = os.path.join(base_dir, "锻材")

    # 如果锻材文件夹不存在，则创建
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # 将YAML字符串写入锻材文件夹中的文件
    with open(os.path.join(folder_path, f"{row[1]}.yml"), "w", encoding='utf-8') as f:
        f.write(yaml_str)
print("YAML文件已成功创建。")
